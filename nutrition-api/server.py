from __future__ import annotations

import json
import os
import re
import threading
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from dotenv import load_dotenv
from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import Workbook, load_workbook

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_FILE = DATA_DIR / "nutrition-db.json"
CACHE_FILE = DATA_DIR / "nutrition-cache.json"
ENTRY_FILE = DATA_DIR / "meal_log.xlsx"
ENTRY_SHEET = "Meals"
ENTRY_HEADERS = [
    "id",
    "date",
    "time",
    "meal_type",
    "food_name",
    "grams",
    "calories",
    "protein",
    "fiber",
    "source",
    "matched_food",
    "image_name",
    "created_at",
]

PORT = int(os.getenv("PORT", "8787"))
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4.1").strip()
OPENAI_API_BASE = os.getenv("OPENAI_API_BASE", "https://api.openai.com/v1").strip().rstrip("/")

app = Flask(__name__)
CORS(app)

local_foods: List[Dict[str, Any]] = []
food_lookup: List[Dict[str, Any]] = []
cache_by_food: Dict[str, Dict[str, Any]] = {}
entry_lock = threading.RLock()


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def round1(value: float) -> float:
    return round(float(value or 0), 1)


def normalize_name(value: str) -> str:
    text = str(value or "").lower()
    text = re.sub(r"[()]", " ", text)
    text = re.sub(r"[^a-z0-9+]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokenize(value: str) -> Set[str]:
    clean = normalize_name(value)
    return {token for token in clean.split(" ") if token}


def jaccard_similarity(tokens_a: Set[str], tokens_b: Set[str]) -> float:
    union = tokens_a | tokens_b
    if not union:
        return 0.0
    return len(tokens_a & tokens_b) / len(union)


def ensure_entry_workbook() -> None:
    if ENTRY_FILE.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = ENTRY_SHEET
    sheet.append(ENTRY_HEADERS)
    workbook.save(ENTRY_FILE)
    workbook.close()


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _row_to_entry(row: Tuple[Any, ...]) -> Optional[Dict[str, Any]]:
    if not row or all(cell is None for cell in row):
        return None

    row_data = list(row)
    if len(row_data) < len(ENTRY_HEADERS):
        row_data.extend([None] * (len(ENTRY_HEADERS) - len(row_data)))

    mapped = dict(zip(ENTRY_HEADERS, row_data))
    entry_id = _to_int(mapped.get("id"), 0)
    if entry_id <= 0:
        return None

    return {
        "id": entry_id,
        "date": str(mapped.get("date") or ""),
        "time": str(mapped.get("time") or ""),
        "mealType": str(mapped.get("meal_type") or "Other"),
        "foodName": str(mapped.get("food_name") or ""),
        "grams": round1(_to_float(mapped.get("grams"))),
        "calories": round1(_to_float(mapped.get("calories"))),
        "protein": round1(_to_float(mapped.get("protein"))),
        "fiber": round1(_to_float(mapped.get("fiber"))),
        "source": str(mapped.get("source") or "unknown"),
        "matchedFood": str(mapped.get("matched_food") or ""),
        "imageName": str(mapped.get("image_name") or ""),
        "createdAt": str(mapped.get("created_at") or ""),
    }


def _entry_to_row(entry: Dict[str, Any]) -> List[Any]:
    return [
        entry.get("id"),
        entry.get("date", ""),
        entry.get("time", ""),
        entry.get("mealType", "Other"),
        entry.get("foodName", ""),
        round1(_to_float(entry.get("grams"))),
        round1(_to_float(entry.get("calories"))),
        round1(_to_float(entry.get("protein"))),
        round1(_to_float(entry.get("fiber"))),
        entry.get("source", "unknown"),
        entry.get("matchedFood", ""),
        entry.get("imageName", ""),
        entry.get("createdAt", ""),
    ]


def list_entries(date_filter: Optional[str] = None) -> List[Dict[str, Any]]:
    with entry_lock:
        ensure_entry_workbook()
        workbook = load_workbook(ENTRY_FILE, data_only=True)
        try:
            if ENTRY_SHEET not in workbook.sheetnames:
                return []

            sheet = workbook[ENTRY_SHEET]
            entries: List[Dict[str, Any]] = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                entry = _row_to_entry(row)
                if not entry:
                    continue
                if date_filter and entry["date"] != date_filter:
                    continue
                entries.append(entry)

            entries.sort(key=lambda item: (item["date"], item["time"], item["id"]))
            return entries
        finally:
            workbook.close()


def write_entries(entries: List[Dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = ENTRY_SHEET
    sheet.append(ENTRY_HEADERS)

    sorted_entries = sorted(entries, key=lambda item: (item["date"], item["time"], item["id"]))
    for entry in sorted_entries:
        sheet.append(_entry_to_row(entry))

    workbook.save(ENTRY_FILE)
    workbook.close()


def validate_date(value: str) -> str:
    text = str(value or "").strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        raise ValueError("date must be in YYYY-MM-DD format.")
    return text


def validate_time(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if not re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", text):
        raise ValueError("time must be in HH:MM 24-hour format.")
    return text


def validate_entry_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    date = validate_date(payload.get("date"))
    time = validate_time(payload.get("time"))

    food_name = str(payload.get("foodName", "")).strip()
    if not food_name:
        raise ValueError("foodName is required.")

    meal_type = str(payload.get("mealType", "Other") or "Other").strip()
    image_name = str(payload.get("imageName", "") or "").strip()
    matched_food = str(payload.get("matchedFood", food_name) or food_name).strip()
    source = str(payload.get("source", "unknown") or "unknown").strip().lower()

    grams = _to_float(payload.get("grams"), -1)
    calories = _to_float(payload.get("calories"), -1)
    protein = _to_float(payload.get("protein"), -1)
    fiber = _to_float(payload.get("fiber"), 0)

    if grams <= 0 or grams > 3000:
        raise ValueError("grams must be between 1 and 3000.")
    if calories < 0 or calories > 5000:
        raise ValueError("calories must be between 0 and 5000.")
    if protein < 0 or protein > 500:
        raise ValueError("protein must be between 0 and 500.")
    if fiber < 0 or fiber > 250:
        raise ValueError("fiber must be between 0 and 250.")

    return {
        "date": date,
        "time": time,
        "mealType": meal_type or "Other",
        "foodName": food_name,
        "grams": round1(grams),
        "calories": round1(calories),
        "protein": round1(protein),
        "fiber": round1(fiber),
        "source": source,
        "matchedFood": matched_food,
        "imageName": image_name,
    }


def create_entry(entry_payload: Dict[str, Any]) -> Dict[str, Any]:
    with entry_lock:
        entries = list_entries(None)
        next_id = max((entry["id"] for entry in entries), default=0) + 1

        created_entry = {
            "id": next_id,
            **entry_payload,
            "createdAt": now_utc_iso(),
        }

        entries.append(created_entry)
        write_entries(entries)
        return created_entry


def remove_entry(entry_id: int) -> bool:
    with entry_lock:
        entries = list_entries(None)
        filtered = [entry for entry in entries if entry["id"] != entry_id]
        if len(filtered) == len(entries):
            return False

        write_entries(filtered)
        return True


def clear_entries_for_date(date: str) -> int:
    with entry_lock:
        entries = list_entries(None)
        filtered = [entry for entry in entries if entry["date"] != date]
        deleted = len(entries) - len(filtered)
        if deleted:
            write_entries(filtered)
        return deleted


def build_lookup(foods: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []

    for food in foods:
        aliases = [food.get("name", "")]
        aliases.extend(food.get("aliases") or [])

        normalized_aliases = sorted(
            {
                normalize_name(alias)
                for alias in aliases
                if isinstance(alias, str) and normalize_name(alias)
            }
        )

        records.append(
            {
                "name": str(food.get("name", "")).strip(),
                "aliases": normalized_aliases,
                "calories": float(food.get("calories", 0)),
                "protein": float(food.get("protein", 0)),
                "fiber": float(food.get("fiber", 0)),
            }
        )

    return records


def compute_match_score(query: str, alias: str) -> float:
    if not query or not alias:
        return 0.0
    if query == alias:
        return 1.0

    overlap = jaccard_similarity(tokenize(query), tokenize(alias))

    if query.startswith(alias) or alias.startswith(query):
        return max(0.82, overlap)
    if query in alias or alias in query:
        return max(0.74, overlap)

    return overlap


def find_local_food(food_name: str) -> Optional[Dict[str, Any]]:
    query = normalize_name(food_name)
    if not query:
        return None

    best: Optional[Dict[str, Any]] = None

    for food in food_lookup:
        for alias in food["aliases"]:
            score = compute_match_score(query, alias)
            if best is None or score > best["score"]:
                best = {"food": food, "alias": alias, "score": score}

    if best is None or best["score"] < 0.7:
        return None

    result = dict(best["food"])
    result["matchedAlias"] = best["alias"]
    result["score"] = best["score"]
    return result


def scale_nutrition(per100: Dict[str, Any], grams: float) -> Dict[str, float]:
    factor = grams / 100.0
    return {
        "calories": round1(float(per100["calories"]) * factor),
        "protein": round1(float(per100["protein"]) * factor),
        "fiber": round1(float(per100["fiber"]) * factor),
    }


def validate_input(payload: Dict[str, Any]) -> Tuple[str, float, bool, bool]:
    food_name = str(payload.get("foodName", "")).strip()
    grams = payload.get("grams")
    allow_ai = bool(payload.get("allowAI", True))
    prefer_ai = bool(payload.get("preferAI", False))

    try:
        grams_value = float(grams)
    except (TypeError, ValueError):
        raise ValueError("grams must be a number between 1 and 3000.")

    if not food_name:
        raise ValueError("foodName is required.")
    if grams_value <= 0 or grams_value > 3000:
        raise ValueError("grams must be a number between 1 and 3000.")
    if prefer_ai and not allow_ai:
        raise ValueError("preferAI requires allowAI=true.")

    return food_name, grams_value, allow_ai, prefer_ai


def load_local_foods() -> None:
    global local_foods, food_lookup

    content = DB_FILE.read_text(encoding="utf-8")
    parsed = json.loads(content)
    local_foods = parsed.get("foods", []) if isinstance(parsed, dict) else []
    food_lookup = build_lookup(local_foods)


def load_cache() -> None:
    global cache_by_food

    if not CACHE_FILE.exists():
        cache_by_food = {}
        return

    content = CACHE_FILE.read_text(encoding="utf-8")
    parsed = json.loads(content) if content.strip() else {}
    cache_by_food = parsed if isinstance(parsed, dict) else {}


def persist_cache() -> None:
    CACHE_FILE.write_text(json.dumps(cache_by_food, indent=2) + "\n", encoding="utf-8")


def extract_json_object(text: str) -> str:
    source = str(text or "").strip()
    if not source:
        return ""

    if source.startswith("{"):
        return source

    fenced = re.search(r"```json\s*([\s\S]*?)```", source, flags=re.IGNORECASE)
    if fenced:
        return fenced.group(1).strip()

    generic_fenced = re.search(r"```\s*([\s\S]*?)```", source)
    if generic_fenced:
        return generic_fenced.group(1).strip()

    object_match = re.search(r"\{[\s\S]*\}", source)
    return object_match.group(0) if object_match else ""


def validate_ai_nutrition(parsed: Dict[str, Any]) -> Dict[str, Any]:
    name = str(parsed.get("name") or parsed.get("food") or "unknown food").strip()

    calories = float(parsed.get("calories"))
    protein = float(parsed.get("protein"))
    fiber = float(parsed.get("fiber"))

    if calories < 0 or calories > 900:
        raise ValueError("Invalid calories estimate returned from AI.")
    if protein < 0 or protein > 100:
        raise ValueError("Invalid protein estimate returned from AI.")
    if fiber < 0 or fiber > 100:
        raise ValueError("Invalid fiber estimate returned from AI.")

    return {
        "name": name,
        "calories": round1(calories),
        "protein": round1(protein),
        "fiber": round1(fiber),
    }


def estimate_with_openai(food_name: str) -> Dict[str, Any]:
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY is missing.")

    system_prompt = "\n".join(
        [
            "You are a nutrition estimation API.",
            "Return nutrition values PER 100 grams of edible portion.",
            "Respond with strict JSON only in this exact schema:",
            '{"name":"<normalized food>","calories":number,"protein":number,"fiber":number}',
            "No explanation text. No markdown.",
        ]
    )

    request_body = {
        "model": OPENAI_MODEL,
        "temperature": 0.1,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": f"Food: {food_name}. Return per-100g calories, protein(g), and fiber(g).",
            },
        ],
    }

    req = urllib.request.Request(
        f"{OPENAI_API_BASE}/chat/completions",
        data=json.dumps(request_body).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {OPENAI_API_KEY}",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=40) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as error:
        details = error.read().decode("utf-8") if error.fp else ""
        try:
            parsed_error = json.loads(details)
            message = parsed_error.get("error", {}).get("message") or details
        except json.JSONDecodeError:
            message = details or str(error)
        raise RuntimeError(message or "OpenAI request failed.") from error
    except urllib.error.URLError as error:
        raise RuntimeError(f"Failed to reach OpenAI API: {error.reason}") from error

    payload = json.loads(raw)
    content = payload.get("choices", [{}])[0].get("message", {}).get("content", "")

    extracted = extract_json_object(content)
    if not extracted:
        raise RuntimeError("AI response did not contain valid JSON.")

    parsed = json.loads(extracted)
    return validate_ai_nutrition(parsed)


def estimate_from_record(record: Dict[str, Any], grams: float, source: str) -> Dict[str, Any]:
    scaled = scale_nutrition(record, grams)
    return {
        **scaled,
        "matchedFood": record["name"],
        "source": source,
        "per100g": {
            "calories": record["calories"],
            "protein": record["protein"],
            "fiber": record["fiber"],
        },
    }


@app.get("/api/health")
def health() -> Any:
    return jsonify(
        {
            "ok": True,
            "model": OPENAI_MODEL,
            "localFoodCount": len(food_lookup),
            "cacheSize": len(cache_by_food),
            "savedMealCount": len(list_entries(None)),
            "aiEnabled": bool(OPENAI_API_KEY),
            "entryDatabase": str(ENTRY_FILE),
        }
    )


@app.get("/api/local-foods")
def local_foods_route() -> Any:
    names = sorted(food["name"] for food in food_lookup)
    return jsonify({"count": len(names), "foods": names})


@app.get("/api/entries")
def list_entries_route() -> Any:
    date = request.args.get("date", "").strip()
    if date:
        try:
            date = validate_date(date)
        except ValueError as error:
            return jsonify({"error": str(error)}), 400

    entries = list_entries(date if date else None)
    return jsonify({"count": len(entries), "entries": entries})


@app.post("/api/entries")
def create_entry_route() -> Any:
    payload = request.get_json(silent=True) or {}

    try:
        validated = validate_entry_payload(payload)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    created = create_entry(validated)
    return jsonify(created), 201


@app.delete("/api/entries/<int:entry_id>")
def delete_entry_route(entry_id: int) -> Any:
    removed = remove_entry(entry_id)
    if not removed:
        return jsonify({"error": "Entry not found."}), 404
    return jsonify({"ok": True, "deleted": 1, "id": entry_id})


@app.delete("/api/entries")
def clear_entries_date_route() -> Any:
    date = request.args.get("date", "").strip()
    if not date:
        return jsonify({"error": "date query parameter is required."}), 400

    try:
        date = validate_date(date)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    deleted = clear_entries_for_date(date)
    return jsonify({"ok": True, "deleted": deleted, "date": date})


@app.post("/api/nutrition/estimate")
def estimate_route() -> Any:
    payload = request.get_json(silent=True) or {}

    try:
        food_name, grams, allow_ai, prefer_ai = validate_input(payload)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    cache_key = normalize_name(food_name)
    cached = cache_by_food.get(cache_key)

    if prefer_ai:
        if cached and cached.get("model") == OPENAI_MODEL:
            return jsonify(estimate_from_record(cached, grams, "cache"))

        try:
            ai_estimate = estimate_with_openai(food_name)
        except Exception as error:  # noqa: BLE001
            return jsonify({"error": f"Unable to estimate food with GPT-4: {error}"}), 502

        cache_by_food[cache_key] = {
            **ai_estimate,
            "cachedAt": now_utc_iso(),
            "model": OPENAI_MODEL,
        }
        persist_cache()
        return jsonify(estimate_from_record(ai_estimate, grams, "openai"))

    local_match = find_local_food(food_name)
    if local_match:
        return jsonify(estimate_from_record(local_match, grams, "local-db"))

    if cached:
        return jsonify(estimate_from_record(cached, grams, "cache"))

    if not allow_ai:
        return (
            jsonify(
                {
                    "error": "Food not found in local database. Enable GPT fallback, or add this food to nutrition-api/data/nutrition-db.json."
                }
            ),
            404,
        )

    try:
        ai_estimate = estimate_with_openai(food_name)
    except Exception as error:  # noqa: BLE001
        return jsonify({"error": f"Unable to estimate unknown food with AI: {error}"}), 502

    cache_by_food[cache_key] = {
        **ai_estimate,
        "cachedAt": now_utc_iso(),
        "model": OPENAI_MODEL,
    }
    persist_cache()

    return jsonify(estimate_from_record(ai_estimate, grams, "openai"))


def bootstrap() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    load_local_foods()
    load_cache()
    if not CACHE_FILE.exists():
        persist_cache()
    ensure_entry_workbook()


bootstrap()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT, debug=True)
